import random
from openpyxl import load_workbook
import sqlite3


file = 'Приложение_к_заданию_бек_разработчика.xlsx'
tablename = 'origin'
tablesorted = 'sorted'
db_row = [
    {
    'id': 'INTEGER',
    'company': 'TEXT',
    'fact_Qliq_data1': 'INTEGER',
    'fact_Qliq_data2': 'INTEGER',
    'fact_Qoil_data1': 'INTEGER',
    'fact_Qoil_data2': 'INTEGER',
    'forecast_Qliq_data1': 'INTEGER',
    'forecast_Qliq_data2': 'INTEGER',
    'forecast_Qoil_data1': 'INTEGER',
    'forecast_Qoil_data2': 'INTEGER',
    'date': 'TEXT'}]
db_row_group = [
    {
    'date': 'TEXT',
    'total_Qliq_data1': 'INTEGER',
    'total_Qliq_data2': 'INTEGER',
    'total_Qoil_data1': 'INTEGER',
    'total_Qoil_data2': 'INTEGER'
    }
]


class DataFile:
    def Read(self):
        raise NotImplementedError()
    
class DBase:
    def addTable(self):
        raise NotImplementedError()
    def addData(self):
        raise NotImplementedError()    

class ReadFile(DataFile):
    def Read(filename: str):
        try:    
            sheet = load_workbook(filename, data_only=True).active
            return sheet
        except:
            return 'error' 
        

class DBaseTable(DBase):
    def addTable(tablename, db_row):
        db_row_add = ', '.join([f'{key} {value}' for key, value in db_row[0].items()])
        cursor = db.cursor()
        query_add_table = f""" CREATE TABLE IF NOT EXISTS {tablename}({db_row_add}) """
        cursor.execute(query_add_table)
        return db_row_add
    
class DBaseDataAdd(DBase):
    def addData(sheet):
            cursor = db.cursor()
            for row in sheet.iter_rows(min_row=4, min_col=1, max_col=10, max_row=sheet.max_row):
                cell = ''
                for i in range(len(row)):
                    if type(row[i].value) == int:
                        cell += f'{row[i].value}, '
                    else:
                        cell += f"'{row[i].value}', "
                cell += f"'{random.randrange(10, 30)}-04-2023'"
                query_add = f""" INSERT INTO {tablename} VALUES({cell}) """
                cursor.execute(query_add)
            
class DBaseSorted(DBase):
    def addData():
        cursor = db.cursor()
        query = f""" SELECT date, SUM(fact_Qliq_data1 - forecast_Qliq_data1) AS total_Qliq_data1,
                                  SUM(fact_Qliq_data2 - forecast_Qliq_data2) AS total_Qliq_data2,
                                  SUM(fact_Qoil_data1 - forecast_Qoil_data1) AS total_Qoil_data1,
                                  SUM(fact_Qoil_data2 - forecast_Qoil_data2) AS total_Qoil_data2
                                  FROM {tablename} GROUP BY date  """
        cursor.execute(query)
        record = cursor.fetchall() 
        cell = ''
        for i in range(len(record)):
            cell = f'{record[i]}, '
            query_add = f""" INSERT INTO {tablesorted} VALUES{cell[:-2]} """
            cursor.execute(query_add)  
 

with sqlite3.connect('database.db') as db:
    try:
        DBaseTable.addTable(tablename = tablename, db_row = db_row)
    except:
        print('Не могу создать таблицу')
    sheet = ReadFile.Read(filename=file) if ReadFile.Read(filename=file) != 'error' else print('Не могу открыть файл')        
    try:    
        DBaseDataAdd.addData(sheet)
    except:
        print('Не могу записать исходные данные')
    try: 
        DBaseTable.addTable(tablename = tablesorted, db_row = db_row_group)
    except:
        print('Не могу создать таблицу с отсортированными данными')  
    try:      
        DBaseSorted.addData()
    except:
        print('Ошибка сортировки данных')    

       



    

